import typing as t

from openpyxl import load_workbook

import definitions


def update_cell_in_xls_file(xlsx_file: str, workbook: str, row: str, column: str, value: str) -> str:
    filepath = definitions.files_resources_folder / xlsx_file
    wb = load_workbook(str(filepath))
    c = wb[workbook].cell(row=row, column=column)
    c.value = value
    wb.save(filepath)
    return xlsx_file


def validate_cell_content(xlsx_file: str, workbook: str, row: str, column: str, expected_value: str, contains: t.Optional[bool] = False) -> None:
    filepath = definitions.resources_folder / xlsx_file
    cell = load_workbook(str(filepath))[workbook].cell(row=row, column=column)

    error_msg = f"XLSX cell verification failed. \nExpected {expected_value}, but was {cell.value}. \nContains: {contains}"
    if contains:
        assert expected_value in cell.value, error_msg
    else:
        assert cell.value == expected_value, error_msg
